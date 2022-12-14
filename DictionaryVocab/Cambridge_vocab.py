import requests, sys
from lxml import etree
from openpyxl import load_workbook
import pandas as pd
import warnings
warnings.filterwarnings("ignore")
import os

def Inputword(word):       
    input_word = input(word)
    return input_word

word = str(Inputword("Vocab: "))
print(word)
def Dictionary(word):
    request = requests.get('https://dictionary.cambridge.org/vi/dictionary/english-chinese-traditional/' + word,
     headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36'})
    
    # parse html
    html = request.text
    page = etree.HTML(html)
    
    # Get definition
    w_definition = page.xpath(
        '//*[@id="page-content"]/div[2]/div[4]/div/div/div[1]/div[3]/div/div[2]/div[1]/div[2]/div'
        )
    if len(w_definition) == 0:
        w_definition = page.xpath(
        '//*[@id="page-content"]/div[2]/div/div[2]/div/div[3]/div/div/div/div[3]/div/div[2]/div[1]/div[2]/div'
        )
    definition_str = w_definition[0].xpath('string(.)').strip()
    print(f'Definition: {definition_str}\n')

    # Get parts of speech
    w_type = page.xpath(
        '//*[@id="page-content"]/div[2]/div[4]/div/div/div/div[2]/div[2]/span[1]'
    )
    type_str = w_type[0].xpath('string(.)').strip()
    
    # Get pronounciation
    w_pronoun = page.xpath(
        '//*[@id="page-content"]/div[2]/div[4]/div/div/div[1]/div[2]/span[2]/span[3]/span[1]'
        )
    if len(w_pronoun) == 0:
        w_pronoun = page.xpath(
    '//*[@id="page-content"]/div[2]/div[4]/div/div/div/div[2]/span[3]/span[3]/span'
    )
    if len(w_pronoun) != 0:
        pronoun_str = w_pronoun[0].xpath('string(.)').strip()
    else:
         pronoun_str = "--"
    text = f'({type_str})   /{pronoun_str}/'
    print(text.center(50,' '))


    isstore = input('Would you like to save in excel workbook? yes or no: ')
    
    if isstore=='yes':        
        file_name = input("Set your Vocab list: ")
        # print(f'1 -- {file_name}')
        StoreExcel(word,type_str, pronoun_str,definition_str,file_name)


def CreateWorkbook(file_name):
    df = pd.DataFrame()
    file = f'{file_name}.xlsx'
    # print(f'2 -- {file}')
    writer= pd.ExcelWriter(file,
                        engine='xlsxwriter')
    df.to_excel(writer,sheet_name = 'Sheet')
    writer.save()
    writer.close()

    return file_name, file

def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def StoreExcel(word,type_str, pronoun_str,definition_str, file_name):
    file = f'{file_name}.xlsx'
    # print(f'3 -- {file}')
    if not os.path.isfile(file):
        file_name, file = CreateWorkbook(file_name)  
        # print(f'4 -- {file}')
        wb = load_workbook(f'{file_name}.xlsx')
        ws = wb.sheetnames
        sheet = wb[ws[0]]
    else:
        wb = load_workbook(f'{file_name}.xlsx')
        ws = wb.sheetnames
        sheet = wb[ws[0]]
    currentCount = get_maximum_rows(sheet_object = sheet)  #.max_row
    # print(f'5 -- {file}')

    # get position
    word_position = 'A' + str(currentCount+1)
    type_position = 'B' + str(currentCount+1)
    pronoun_position = 'C' + str(currentCount+1)
    meaning_position = 'D' + str(currentCount+1)

    # write new word
    sheet[word_position] = word
    sheet[type_position] = type_str
    sheet[pronoun_position] = pronoun_str
    sheet[meaning_position] = definition_str    
    
    afterCount = get_maximum_rows(sheet_object = sheet)
    
    print('Saved Word!')
    print(f'{file} has {afterCount} words.')
    
    # save file
    wb.save(file)


    
if len(word) > 1:
    Dictionary(word)
else:
    print('No word!')

